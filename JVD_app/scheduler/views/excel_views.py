# Excel imports
import xlsxwriter
from django.contrib.staticfiles import finders

from calendar import monthrange
from django.db.models import Sum
from django.db.models import IntegerField, Value
from django.db.models.functions import Cast, Coalesce
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from datetime import date, timedelta, datetime
from ..models import ShiftType, Employee, WorkDay, FixedHourFund, Holiday, ExcessHours
from io import BytesIO
import calendar
from django.contrib.auth.decorators import login_required
import logging
logger = logging.getLogger(__name__)

import openpyxl
from datetime import timedelta, datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.styles import Font

from django.db.models import Case, When, IntegerField, Value
from django.db.models.functions import Coalesce


#--------------------------------------------------------------------------
#--------------------- Generiranje Excel datoteki -------------------------
#--------------------------------------------------------------------------

def is_night_shift_start(wd):
    if not wd.shift_type.isNightShift:
        return False
    if wd.shift_type.name == 'Priprema od 19:00':
        # Check if it's the starting day based on on_call_hours
        if wd.on_call_hours == 5:
            return True  # Starting day
        else:
            return False  # Next day
    # Existing logic for other night shifts
    if wd.day_hours >= 3:
        return True
    else:
        return False



def get_week_dates(start_date):
    # start_date is assumed to be a Monday
    return [start_date + timedelta(days=i) for i in range(7)]


# Croatian day names to use for mapping
day_names_cro = {
    'Monday': 'Pon', 
    'Tuesday': 'Uto', 
    'Wednesday': 'Sri', 
    'Thursday': 'Čet', 
    'Friday': 'Pet', 
    'Saturday': 'Sub', 
    'Sunday': 'Ned'
}

def croatian_day(date_obj):
    """Return the Croatian day abbreviation for the date."""
    return day_names_cro[calendar.day_name[date_obj.weekday()]]


def get_date_table_mapping(start_date, total_days, table_start_rows):
    """
    Returns a mapping from date strings to their corresponding table index and starting row.
    """
    date_table_mapping = {}
    current_date = start_date
    day_counter = 0

    for table_index, table_start_row in enumerate(table_start_rows):
        for day_in_week in range(7):  # Each table has 7 days
            if day_counter >= total_days:
                break  # Stop if we've covered all days

            date_str = current_date.strftime('%Y-%m-%d')
            date_table_mapping[date_str] = {
                'table_index': table_index,
                'table_start_row': table_start_row,
                'day_offset': day_in_week
            }

            current_date += timedelta(days=1)
            day_counter += 1

    return date_table_mapping


def fill_dates_in_template(wb, date_table_mapping):
    sheet = wb.active

    for date_str, mapping in date_table_mapping.items():
        table_start_row = mapping['table_start_row']
        day_offset = mapping['day_offset']

        # Calculate the starting row for this date within the table
        day_start_row = table_start_row + day_offset * 7

        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()

        # Fill date information in Column B
        sheet[f"B{day_start_row}"] = croatian_day(date_obj)  # Day abbreviation
        sheet[f"B{day_start_row + 1}"] = date_obj.strftime('%d/%m')  # Day/Month
        sheet[f"B{day_start_row + 2}"] = date_obj.year  # Year


def fill_employees_in_template(wb, schedule_data, date_table_mapping):
    sheet = wb.active

    # Updated shift_columns mapping
    shift_columns = {
        '1. Smjena': 'C',
        '2.smjena (Priprema od 7:00)': 'D',
        'Priprema od 19:00': 'E',
        'JANAF 1.smjena': 'F',
        'JANAF 2.Smjena': 'F',  # Both shifts use column F
        'Slobodan Dan 1': 'G',
        'Slobodan Dan 2': 'H',
        'Godišnji odmor': 'I',
        'Bolovanje': 'J',
        'INA 1.Smjena': 'K',
        'INA 2. Smjena': 'L',
    }

    # Group colors mapping for font color
    group_colors = {
        '2': 'FF0000',  # Red
        '3': '008000',  # Green
        '4': '0000FF',  # Blue
        '5': '000000',  # Black
        '6': '800080',  # Purple
    }

    # Starting rows for special shifts in each table
    special_shift_rows = {
        0: {'JANAF 1.smjena': 6, 'JANAF 2.Smjena': 9, 'Slobodan Dan 1': 5, 'Slobodan Dan 2': 5, 'Other Shifts': 7},
        1: {'JANAF 1.smjena': 59, 'JANAF 2.Smjena': 62, 'Slobodan Dan 1': 58, 'Slobodan Dan 2': 58, 'Other Shifts': 60},
        2: {'JANAF 1.smjena': 112, 'JANAF 2.Smjena': 115, 'Slobodan Dan 1': 111, 'Slobodan Dan 2': 111, 'Other Shifts': 113},
        3: {'JANAF 1.smjena': 166, 'JANAF 2.Smjena': 169, 'Slobodan Dan 1': 165, 'Slobodan Dan 2': 165, 'Other Shifts': 167},
        4: {'JANAF 1.smjena': 219, 'JANAF 2.Smjena': 222, 'Slobodan Dan 1': 218, 'Slobodan Dan 2': 218, 'Other Shifts': 220},
        5: {'JANAF 1.smjena': 272, 'JANAF 2.Smjena': 275, 'Slobodan Dan 1': 271, 'Slobodan Dan 2': 271, 'Other Shifts': 273},
    }

    for date_str, day_data in schedule_data.items():
        if date_str in date_table_mapping:
            mapping = date_table_mapping[date_str]
            table_index = mapping['table_index']
            day_offset = mapping['day_offset']

            # Get starting rows for this table
            table_shifts = special_shift_rows[table_index]

            for shift_type_id, employees_list in day_data.items():
                if employees_list:
                    shift_type = ShiftType.objects.get(id=shift_type_id)
                    shift_name = shift_type.name
                    shift_column = shift_columns.get(shift_name, None)

                    if not shift_column:
                        continue  # Skip if shift column isn't defined

                    if shift_name in ['JANAF 1.smjena', 'JANAF 2.Smjena']:
                        # Get starting row for the shift in the table
                        shift_start_row = table_shifts[shift_name]

                        # Adjust for day offset
                        shift_start_row += day_offset * 7

                        if shift_name == 'JANAF 1.smjena':
                            rows = [shift_start_row, shift_start_row + 1]
                        else:
                            rows = [shift_start_row, shift_start_row + 1, shift_start_row + 2]

                        for idx, employee in enumerate(employees_list):
                            if idx < len(rows):
                                cell = sheet[f"{shift_column}{rows[idx]}"]
                                # Add role number to the name
                                cell.value = f"{employee.surname} {employee.name[0]}. ({employee.role_number})"

                                # Get the group value and color
                                group_value = str(employee.group)
                                color_code = group_colors.get(group_value, '000000')  # Default to black

                                # Apply the font color
                                font = Font(name='Times New Roman', color=color_code, size=7, bold=True)
                                cell.font = font

                    elif shift_name in ['Slobodan Dan 1', 'Slobodan Dan 2']:
                        # Get starting row for the shift in the table
                        shift_start_row = table_shifts[shift_name]

                        # Adjust for day offset
                        shift_start_row += day_offset * 7

                        rows = [shift_start_row + i for i in range(7)]  # 7 rows per day

                        for idx, employee in enumerate(employees_list):
                            if idx < len(rows):
                                cell = sheet[f"{shift_column}{rows[idx]}"]
                                # Add role number to the name
                                cell.value = f"{employee.surname} {employee.name[0]}. ({employee.role_number})"

                                # Get the group value and color
                                group_value = str(employee.group)
                                color_code = group_colors.get(group_value, '000000')  # Default to black

                                # Apply the font color
                                font = Font(name='Times New Roman', color=color_code, size=7, bold=True)
                                cell.font = font

                    else:
                        # Other shifts
                        shift_start_row = table_shifts['Other Shifts']

                        # Adjust for day offset
                        shift_start_row += day_offset * 7

                        for idx, employee in enumerate(employees_list):
                            cell = sheet[f"{shift_column}{shift_start_row + idx}"]
                            # Add role number to the name
                            cell.value = f"{employee.surname} {employee.name[0]}. ({employee.role_number})"

                            # Get the group value and color
                            group_value = str(employee.group)
                            color_code = group_colors.get(group_value, '000000')  # Default to black

                            # Apply the font color
                            font = Font(name='Times New Roman', color=color_code, size=7, bold=True)
                            cell.font = font



def create_schedule_excel(week_dates, shift_types, schedule_data, author_name):
    # Use Django staticfiles finders to locate the Excel template
    template_path = finders.find('template/RASPORED_template.xlsx')
    
    if not template_path:
        raise FileNotFoundError("The Excel template file was not found.")
    
    # Load the Excel workbook from the template
    wb = openpyxl.load_workbook(template_path)
    
    # Define table starting rows
    table_start_rows = [7, 60, 113, 167, 220, 273]  # Starting rows for each table

    # Calculate the total number of days to cover
    total_days = (week_dates[-1] - week_dates[0]).days + 1

    # Generate date-to-table mapping
    date_table_mapping = get_date_table_mapping(week_dates[0], total_days, table_start_rows)

    # Fill the dates in the template
    fill_dates_in_template(wb, date_table_mapping)

    # Fill employees in the template
    fill_employees_in_template(wb, schedule_data, date_table_mapping)

    # Save the modified workbook to an in-memory stream (BytesIO)
    from io import BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return excel_file



@login_required
def download_schedule(request):
    month_str = request.GET.get('month')
    if month_str:
        start_date = datetime.strptime(month_str, '%Y-%m-%d').date()
    else:
        today = date.today()
        start_date = today - timedelta(days=today.weekday())

    # Adjust start_date to the Monday before the start of the month
    start_date = start_date - timedelta(days=start_date.weekday())

    # Calculate end_date to cover all weeks in the month (6 weeks)
    end_date = start_date + timedelta(days=6 * 7 - 1)

    # Generate list of dates to cover
    total_days = (end_date - start_date).days + 1
    week_dates = [start_date + timedelta(days=i) for i in range(total_days)]

    shift_types = ShiftType.objects.all()

    # Prepare schedule data using WorkDay
    schedule_data = {}
    for day in week_dates:
        day_str = day.strftime('%Y-%m-%d')
        schedule_data[day_str] = {}
        for shift_type in shift_types:
            schedule_data[day_str][shift_type.id] = []

    # Fetch WorkDay entries and populate schedule_data
    workdays = WorkDay.objects.filter(date__in=week_dates).select_related('employee', 'shift_type')
    for wd in workdays:
        include_wd = True
        if wd.shift_type.isNightShift:
            if not is_night_shift_start(wd):
                include_wd = False
        if include_wd:
            day_str = wd.date.strftime('%Y-%m-%d')
            shift_type_id = wd.shift_type.id
            schedule_data[day_str][shift_type_id].append(wd.employee)

    author_name = request.user.username

    excel_file = create_schedule_excel(week_dates, shift_types, schedule_data, author_name)

    response = HttpResponse(content=excel_file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Raspored.xlsx"'
    return response



@login_required
def download_sihterica(request):
    return download_sihterica_with_filters(request, exclude_group='6')

@login_required
def download_sihterica_ina(request):
    return download_sihterica_with_filters(request, group_filter='6')

def download_sihterica_with_filters(request, group_filter=None, exclude_group=None):
    month_str = request.GET.get('month')
    try:
        if month_str:
            start_date = datetime.strptime(month_str, '%Y-%m-%d').date()
        else:
            start_date = date.today().replace(day=1)
        
        current_year = start_date.year
        current_month = start_date.month
        days_in_month = monthrange(current_year, current_month)[1]
        
        # Define a role priority for group 1 employees
        role_priority = Case(
            When(role='Zapovjednik', then=Value(1)),
            When(role='Zamjenik zapovjednika', then=Value(2)),
            When(role='Rukovatelj opcih ekonomskih i komercijalnih poslova', then=Value(3)),
            When(role='Vatrogasac preventivac', then=Value(4)),
            When(role='Spremacica', then=Value(5)),
            When(role='Viši referent za ekonomske poslove', then=Value(6)),
            default=Value(999),
            output_field=IntegerField()
        )

        # For group 1 (office) employees, sort by custom role order, then surname, then name
        group1_employees = Employee.objects.filter(group='1').annotate(
            role_order=role_priority
        ).order_by('role_order', 'surname', 'name')

        # For other employees, sort by group_number, role_number_int, then surname, then name
        other_employees = Employee.objects.exclude(group='1').annotate(
            group_number=Cast('group', output_field=IntegerField()),
            role_number_int=Coalesce('role_number', Value(0))
        ).order_by('group_number', 'role_number_int', 'surname', 'name')

        # Apply filters if provided
        if group_filter:
            group1_employees = group1_employees.filter(group=group_filter)
            other_employees = other_employees.filter(group=group_filter)
        if exclude_group:
            group1_employees = group1_employees.exclude(group=exclude_group)
            other_employees = other_employees.exclude(group=exclude_group)

        # Combine
        employees = list(group1_employees) + list(other_employees)

        # Now employees is correctly filtered and sorted
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})

        # Main timesheet worksheet
        timesheet_worksheet = workbook.add_worksheet('Šihterica')

        # Set up formats
        formats = setup_formats(workbook)
        timesheet_formats = setup_timesheet_formats(workbook)

        write_headers(timesheet_worksheet, current_month, current_year, days_in_month, timesheet_formats)
        fill_timesheet(timesheet_worksheet, employees, current_year, current_month, days_in_month, timesheet_formats)

        # Generate other sheets
        overview_worksheet = workbook.add_worksheet('Pregled svih sati')
        generate_total_overview(overview_worksheet, employees, current_year, current_month, formats)

        preparation_worksheet = workbook.add_worksheet('Sati pripreme')
        generate_preparation_hours(preparation_worksheet, employees, current_year, current_month, formats)

        excess_worksheet = workbook.add_worksheet('Evidencija viška-manjka sati')
        generate_excess_hours(excess_worksheet, employees, current_year, current_month, formats)

        vacation_worksheet = workbook.add_worksheet('Evidencija godišnjih odmora')
        generate_vacation_hours(vacation_worksheet, employees, current_year, current_month, formats)

        overtime_worksheet = workbook.add_worksheet('Evidencija prekovremenih sati')
        generate_overtime_hours(overtime_worksheet, employees, current_year, current_month, formats)

        # Close workbook, produce response
        workbook.close()
        output.seek(0)
        response = HttpResponse(
            content=output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        if exclude_group == '6':
            filename_suffix = ''
        elif group_filter == '6':
            filename_suffix = 'INA '
        else:
            filename_suffix = ''
        response['Content-Disposition'] = (
            f'attachment; filename="{filename_suffix}sihterica_{current_month}_{current_year}.xlsx"'
        )

        return response

    except Exception as e:
        logger.exception("Error generating sihterica (group_filter=%s, exclude_group=%s, month=%s)", group_filter, exclude_group, month_str)
        # You can either return a plain HttpResponse or render an error page.
        return HttpResponse("Greška prilikom generiranja Šihterice.", status=500)


def setup_formats(workbook):
    """Setup cell formats for the workbook."""
    common_format = {'align': 'center', 'valign': 'vcenter', 'font_size': 12, 'text_wrap': True, 'border': 1}
    formats = {
        'title_format': workbook.add_format({**common_format, 'bold': True, 'font_size': 16}),
        'bold_format': workbook.add_format({**common_format, 'bold': True}),
        'header_format': workbook.add_format({**common_format, 'bold': True, 'bg_color': '#f0f0f0'}),
        'total_format': workbook.add_format({**common_format, 'bg_color': '#d9e1f2'}),
        'hours_format': workbook.add_format({**common_format}),
        'saturday_format': workbook.add_format({**common_format, 'bg_color': '#CCFFCC'}),  # Green
        'sunday_format': workbook.add_format({**common_format, 'bg_color': '#FFFF99'}),  # Yellow
        'gray_format': workbook.add_format({**common_format, 'bg_color': '#E0E0E0'}),  # Gray
        'white_format': workbook.add_format({**common_format, 'bg_color': '#FFFFFF'}),  # White
        'vacation_format': workbook.add_format({**common_format, 'font_color': '#00FF00'}),
        'total_vacation_format': workbook.add_format({**common_format, 'font_color': '#00FF00', 'bg_color': '#d9e1f2'})
    }
    # Group text color formats
    formats['group_text_formats'] = {
        '1': workbook.add_format({**common_format, 'color': '#1E8449', 'bold': True}),
        '2': workbook.add_format({**common_format, 'color': '#D35400', 'bold': True}),
        '3': workbook.add_format({**common_format, 'color': '#2980B9', 'bold': True}),
        '4': workbook.add_format({**common_format, 'color': '#2980B9', 'bold': True}),
        '5': workbook.add_format({**common_format, 'color': '#2980B9', 'bold': True}),
        '6': workbook.add_format({**common_format, 'color': '#2980B9', 'bold': True}),
    }
    return formats

def setup_timesheet_formats(workbook, vacation_color='#008000', sick_color='#FF0000', holiday_color='#D35400'):
    """Setup cell formats specifically for the timesheet."""
    # Keeping your original comments:
    # Here we set up the formats for the timesheet including new colors for vacation, sick leave, and holiday
    
    formats = {
        'title_format': workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'font_size': 16}),
        'bold_format': workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'font_size': 12}),
        'header_format': workbook.add_format({'align': 'center', 'bold': True, 'bg_color': '#f0f0f0'}),
        'total_format': workbook.add_format({'align': 'center', 'bg_color': '#d9e1f2'}),
        'hours_format': workbook.add_format({'align': 'center'}),
        'saturday_format': workbook.add_format({'align': 'center', 'bg_color': '#CCFFCC'}),  # Green for Saturday
        'sunday_format': workbook.add_format({'align': 'center', 'bg_color': '#FFFF99'}),   # Yellow for Sunday
        'gray_format': workbook.add_format({'bg_color': '#E0E0E0', 'align': 'center'}),     # Gray
        'white_format': workbook.add_format({'bg_color': '#FFFFFF', 'align': 'center'}),    # White

        # Vacation now more visible green
        'vacation_format': workbook.add_format({'align': 'center', 'font_color': vacation_color}),
        'total_vacation_format': workbook.add_format({'align': 'center', 'font_color': vacation_color, 'bg_color': '#d9e1f2'}),

        # Sick leave red in timesheet
        'sick_leave_format': workbook.add_format({'align': 'center', 'font_color': sick_color}),
        'total_sick_leave_format': workbook.add_format({'align': 'center', 'font_color': sick_color, 'bg_color': '#d9e1f2'}),

        # Holiday brown/orange
        'holiday_format': workbook.add_format({'align': 'center', 'font_color': holiday_color}),
        'total_holiday_format': workbook.add_format({'align': 'center', 'font_color': holiday_color, 'bg_color': '#d9e1f2'})
    }

    # Group text color formats remain the same, keeping original comments:
    # Group text color formats for name formatting:
    formats['group_text_formats'] = {
        '1': workbook.add_format({'color': '#9C7C14', 'align': 'center', 'valign': 'vcenter', 'font_size': 12, 'bold': True}),
        '2': workbook.add_format({'color': '#c9242d', 'align': 'center', 'valign': 'vcenter', 'font_size': 12, 'bold': True}),
        '3': workbook.add_format({'color': '#118f37', 'align': 'center', 'valign': 'vcenter', 'font_size': 12, 'bold': True}),
        '4': workbook.add_format({'color': '#16448a', 'align': 'center', 'valign': 'vcenter', 'font_size': 12, 'bold': True}),
        '5': workbook.add_format({'color': '#000000', 'align': 'center', 'valign': 'vcenter', 'font_size': 12, 'bold': True}),
        '6': workbook.add_format({'color': '#6b157a', 'align': 'center', 'valign': 'vcenter', 'font_size': 12, 'bold': True}),
    }

    return formats


def write_headers(worksheet, current_month, current_year, days_in_month, formats):
    """Write headers for the worksheet based on current month and year."""
    croatian_months = {
        1: 'Siječanj', 2: 'Veljača', 3: 'Ožujak', 4: 'Travanj', 5: 'Svibanj', 6: 'Lipanj',
        7: 'Srpanj', 8: 'Kolovoz', 9: 'Rujan', 10: 'Listopad', 11: 'Studeni', 12: 'Prosinac'
    }
    month_name = croatian_months[current_month]
    title = f"Evidencija prisutnosti na radu za mjesec {month_name} {current_year}"
    worksheet.merge_range(0, 0, 0, 43, title, formats['title_format'])

    # Set column widths
    worksheet.set_column(0, 0, 20)  # Employee names
    worksheet.set_column(1, 1, 3)    # RD/RN column
    worksheet.set_column(2, 33, 4)   # Day columns
    worksheet.set_column(34, 43, 15) # Aggregate columns

    # Add headers for days and aggregates
    worksheet.merge_range(1, 0, 2, 0, 'Prezime i ime', formats['header_format'])
    day_abbreviations = ['P', 'U', 'S', 'Č', 'P', 'S', 'N']
    for day in range(1, days_in_month + 1):
        weekday = date(year=current_year, month=current_month, day=day).weekday()
        format_to_use = formats['saturday_format'] if weekday == 5 else formats['sunday_format'] if weekday == 6 else formats['header_format']
        worksheet.write(1, day + 1, day_abbreviations[weekday], format_to_use)
        worksheet.write(2, day + 1, str(day), format_to_use)

    # Aggregates headers
    aggregates = ['Fond sati', 'Redovan Rad', 'Turnus', 'Višak fonda', 'Subota', 'Nedjelja', 'Noćni rad', 'Blagdan', 'Bolovanje', 'Godišnji odmor']
    for idx, name in enumerate(aggregates, start=34):
        worksheet.merge_range(1, idx, 2, idx, name, formats['total_format'])

def fill_timesheet(worksheet, employees, current_year, current_month, days_in_month, formats):
    row_idx = 3
    for employee in employees:
        group_format = formats['group_text_formats'].get(employee.group, formats['bold_format'])
        worksheet.merge_range(row_idx, 0, row_idx + 1, 0, f"{employee.surname} {employee.name}", group_format)
        worksheet.write(row_idx, 1, 'RD', formats['bold_format'])
        worksheet.write(row_idx + 1, 1, 'RN', formats['bold_format'])

        total_day_hours = total_night_hours = 0
        total_saturday_hours = total_sunday_hours = 0
        total_holiday_hours = total_sick_leave_hours = total_vacation_hours = 0

        for day in range(1, days_in_month + 1):
            current_date = date(year=current_year, month=current_month, day=day)
            work_days = WorkDay.objects.filter(employee=employee, date=current_date)
            weekday = current_date.weekday()
            base_cell_format = (
                formats['saturday_format'] if weekday == 5 else
                formats['sunday_format'] if weekday == 6 else
                formats['hours_format']
            )

            if work_days.exists():
                total_day_hours_for_day = sum((wd.day_hours or 0) for wd in work_days)
                total_night_hours_for_day = sum((wd.night_hours or 0) for wd in work_days)
                total_vacation_hours_for_day = sum((wd.vacation_hours or 0) for wd in work_days)
                total_sick_for_day = sum((wd.sick_leave_hours or 0) for wd in work_days)
                total_holiday_for_day = sum((wd.holiday_hours or 0) for wd in work_days)

                day_hours_display = total_day_hours_for_day if total_day_hours_for_day != 0 else ''
                night_hours_display = total_night_hours_for_day if total_night_hours_for_day != 0 else ''

                if total_vacation_hours_for_day > 0:
                    # Vacation
                    worksheet.write(row_idx, day + 1, '0', formats['vacation_format'])
                    worksheet.write(row_idx + 1, day + 1, night_hours_display, base_cell_format)
                elif total_sick_for_day > 0:
                    # Sick leave: show sick leave hours on RD row and blank on RN
                    worksheet.write(row_idx, day + 1, '0', formats['sick_leave_format'])
                    worksheet.write(row_idx + 1, day + 1, '', formats['sick_leave_format'])
                elif total_holiday_for_day > 0:
                    # Holiday
                    worksheet.write(row_idx, day + 1, day_hours_display if day_hours_display else '', formats['holiday_format'])
                    worksheet.write(row_idx + 1, day + 1, night_hours_display if night_hours_display else '', formats['holiday_format'])
                else:
                    # Normal day/night hours
                    worksheet.write(row_idx, day + 1, day_hours_display, base_cell_format)
                    worksheet.write(row_idx + 1, day + 1, night_hours_display, base_cell_format)

                total_day_hours += total_day_hours_for_day
                total_night_hours += total_night_hours_for_day
                total_vacation_hours += total_vacation_hours_for_day
                total_sick_leave_hours += total_sick_for_day
                total_holiday_hours += total_holiday_for_day

                if weekday == 5:
                    total_saturday_hours += total_day_hours_for_day + total_night_hours_for_day
                elif weekday == 6:
                    total_sunday_hours += total_day_hours_for_day + total_night_hours_for_day
            else:
                worksheet.write(row_idx, day + 1, '', base_cell_format)
                worksheet.write(row_idx + 1, day + 1, '', base_cell_format)

        try:
            hour_fund = FixedHourFund.objects.get(month__year=current_year, month__month=current_month).required_hours
        except FixedHourFund.DoesNotExist:
            hour_fund = 168

        redovan_rad = employee.calculate_redovan_rad(current_month, current_year)
        turnus = employee.calculate_monthly_hours(current_month, current_year)
        visak_sati = employee.calculate_visak_sati(current_month, current_year)

        worksheet.merge_range(row_idx, 34, row_idx + 1, 34, hour_fund, formats['total_format'])
        worksheet.merge_range(row_idx, 35, row_idx + 1, 35, redovan_rad, formats['total_format'])
        worksheet.merge_range(row_idx, 36, row_idx + 1, 36, turnus, formats['total_format'])
        worksheet.merge_range(row_idx, 37, row_idx + 1, 37, visak_sati, formats['total_format'])
        worksheet.merge_range(row_idx, 38, row_idx + 1, 38, total_saturday_hours, formats['total_format'])
        worksheet.merge_range(row_idx, 39, row_idx + 1, 39, total_sunday_hours, formats['total_format'])
        worksheet.merge_range(row_idx, 40, row_idx + 1, 40, total_night_hours, formats['total_format'])
        worksheet.merge_range(row_idx, 41, row_idx + 1, 41, total_holiday_hours, formats['total_holiday_format'])
        worksheet.merge_range(row_idx, 42, row_idx + 1, 42, total_sick_leave_hours, formats['total_sick_leave_format'])
        worksheet.merge_range(row_idx, 43, row_idx + 1, 43, total_vacation_hours, formats['total_vacation_format'])

        row_idx += 2




def generate_total_overview(worksheet, employees, current_year, current_month, formats):
    from datetime import date
    from calendar import monthrange
    import xlsxwriter

    row_idx = 0  # Start at the top of the new worksheet
    agg_start_col = 0  # Start at the first column

    worksheet.set_column(0, 0, 5)  # rb. column
    worksheet.set_column(1, 1, 30)  # Employee name column
    worksheet.set_column(2, 2, 12)  # Fond sati column
    worksheet.set_column(3, 15, 15)

    croatian_months = {
        1: 'Siječanj', 2: 'Veljača', 3: 'Ožujak', 4: 'Travanj', 5: 'Svibanj', 6: 'Lipanj',
        7: 'Srpanj', 8: 'Kolovoz', 9: 'Rujan', 10: 'Listopad', 11: 'Studeni', 12: 'Prosinac'
    }
    month_name = croatian_months[current_month]

    # Title for the first aggregate table
    worksheet.merge_range(row_idx, agg_start_col, row_idx, agg_start_col + 15, f'Ukupni sati za {month_name} {current_year}', formats['title_format'])
    row_idx += 1

    # Add first aggregate table headers
    aggregate_headers = ['rb.', 'Prezime i ime', 'Fond sati', 'Red. rad', 'Drž. pr. i bla.', 'Godišnji o.', 'Bolovanje', 'Noćni rad',
                         'Rad sub.', 'Rad. ned.', 'Slobodan dan', 'Turnus', 'Priprema', 'Prek.rad', 'Prek. USLUGA', 'Prek. Višak Fonda']

    for i, header in enumerate(aggregate_headers):
        worksheet.merge_range(row_idx, agg_start_col + i, row_idx + 1, agg_start_col + i, header, formats['header_format'])

    row_idx += 2

    # Add first aggregate data
    for idx, employee in enumerate(employees):
        work_days = WorkDay.objects.filter(employee=employee, date__year=current_year, date__month=current_month)

        # Handle None values by using 'or 0'
        total_vacation_hours = sum((day.vacation_hours or 0) for day in work_days)
        total_sick_leave_hours = sum((day.sick_leave_hours or 0) for day in work_days)
        total_holiday_hours = sum((day.holiday_hours or 0) for day in work_days)
        total_day_hours = sum((day.day_hours or 0) for day in work_days)
        total_night_hours = sum((day.night_hours or 0) for day in work_days)

        # Calculate Saturday and Sunday hours including night hours
        days_in_month = monthrange(current_year, current_month)[1]
        saturday_hours = 0
        sunday_hours = 0
        for day_num in range(1, days_in_month + 1):
            current_date = date(current_year, current_month, day_num)
            weekday = current_date.weekday()
            day_work_days = work_days.filter(date=current_date)
            day_total_day_hours = sum((wd.day_hours or 0) for wd in day_work_days)
            day_total_night_hours = sum((wd.night_hours or 0) for wd in day_work_days)
            if weekday == 5:  # Saturday
                saturday_hours += day_total_day_hours + day_total_night_hours
            elif weekday == 6:  # Sunday
                sunday_hours += day_total_day_hours + day_total_night_hours

        total_saturday_hours = saturday_hours
        total_sunday_hours = sunday_hours

        # Calculate free days hours (if you have such a field)
        total_free_days_hours = sum((day.article39_hours or 0) for day in work_days)

        # Total on-call hours
        total_on_call_hours = sum((day.on_call_hours or 0) for day in work_days)

        # Total overtime hours
        total_overtime_hours = sum((day.overtime_hours or 0) for day in work_days)
        total_overtime_service = sum((day.overtime_service or 0) for day in work_days)
        total_overtime_excess_fond = sum((day.overtime_excess_fond or 0) for day in work_days)

        # Calculate hour_fund for each employee
        try:
            hour_fund = FixedHourFund.objects.get(month__year=current_year, month__month=current_month).required_hours
        except FixedHourFund.DoesNotExist:
            hour_fund = 168  # Default or error handling scenario

        # Use methods from models.py
        redovan_rad = employee.calculate_redovan_rad(current_month, current_year)
        turnus = employee.calculate_monthly_hours(current_month, current_year)
        visak_sati = employee.calculate_visak_sati(current_month, current_year)

        row_format = formats['gray_format'] if idx % 2 == 0 else formats['white_format']

        values = [
            idx + 1,  # rb.
            f"{employee.surname} {employee.name}",  # Prezime i ime
            hour_fund,  # Fond sati
            redovan_rad,  # Red. rad
            total_holiday_hours,  # Drž. pr. i bla.
            total_vacation_hours,  # Godišnji o.
            total_sick_leave_hours,  # Bolovanje
            total_night_hours,  # Noćni rad
            total_saturday_hours,  # Rad sub.
            total_sunday_hours,  # Rad. ned.
            total_free_days_hours,  # Slobodan dan
            turnus,  # Turnus
            total_on_call_hours,  # Priprema
            total_overtime_hours,  # Prek.rad
            total_overtime_service,  # Prek. USLUGA
            total_overtime_excess_fond,  # Prek. Višak Fonda
        ]


        for col, value in enumerate(values):
            worksheet.write(row_idx, agg_start_col + col, value, row_format)

        row_idx += 1

    # Add "Ukupno" row for the first aggregate table
    first_data_row = row_idx - len(employees)
    worksheet.write(row_idx, agg_start_col+1, 'Ukupno', formats['header_format'])

    # Sum columns from 'Fond sati' (col_num=2) to 'Prek. Višak Fonda' (col_num=15)
    for col_num in range(2, 16):
        column_letter_value = xlsxwriter.utility.xl_col_to_name(col_num)
        formula_range = f"{column_letter_value}{first_data_row + 1}:{column_letter_value}{row_idx}"
        worksheet.write_formula(row_idx, col_num, f"=SUM({formula_range})", formats['total_format'])


def generate_preparation_hours(worksheet, employees, current_year, current_month, formats):
    worksheet.set_column(0, 0, 5)  # rb. column
    worksheet.set_column(1, 1, 30)  # Employee name column
    worksheet.set_column(2, 6, 10)  # Week columns
    worksheet.set_column(7, 7, 20)  # Priprema um. prekovremene column
    worksheet.set_column(8, 8, 12)  # Ukupno column

    first_day_of_month = date(current_year, current_month, 1)
    last_day_of_month = date(current_year, current_month, monthrange(current_year, current_month)[1])
    weeks = {((first_day_of_month + timedelta(days=x)).isocalendar()[1]) for x in range((last_day_of_month - first_day_of_month).days + 1)}

    agg_start_row = 0
    worksheet.merge_range(agg_start_row, 0, agg_start_row, 8, 'BROJ SATI PRIPREME', formats['title_format'])
    agg_start_row += 1

    # Write headers (rb. then Prezime i ime)
    worksheet.write(agg_start_row, 0, 'rb.', formats['header_format'])
    worksheet.write(agg_start_row, 1, 'Prezime i ime', formats['header_format'])

    start_week_col = 2
    end_week_col = start_week_col + len(weeks) - 1
    worksheet.merge_range(agg_start_row, start_week_col, agg_start_row, end_week_col, 'Prip. iz rasporeda', formats['header_format'])
    worksheet.write(agg_start_row, end_week_col + 1, 'Priprema um. prekovremene', formats['header_format'])
    worksheet.write(agg_start_row, end_week_col + 2, 'Ukupno', formats['header_format'])

    agg_start_row += 1

    for idx, employee in enumerate(employees):
        row_format = formats['gray_format'] if idx % 2 == 0 else formats['white_format']
        # Now we can use idx and employee
        worksheet.write(agg_start_row, 0, idx + 1, row_format)  # rb.
        worksheet.write(agg_start_row, 1, f"{employee.surname} {employee.name}", row_format)

        col_offset = 2
        total_weekly_hours = 0
        for week_num in sorted(weeks):
            week_days = [first_day_of_month + timedelta(days=x) 
                         for x in range((last_day_of_month - first_day_of_month).days + 1) 
                         if (first_day_of_month + timedelta(days=x)).isocalendar()[1] == week_num]
            weekly_hours = sum(WorkDay.objects.filter(employee=employee, date__in=week_days).values_list('on_call_hours', flat=True))
            worksheet.write(agg_start_row, col_offset, weekly_hours, row_format)
            total_weekly_hours += weekly_hours
            col_offset += 1

        worksheet.write(agg_start_row, col_offset, 0, row_format)
        worksheet.write(agg_start_row, col_offset + 1, total_weekly_hours, row_format)

        agg_start_row += 1

    total_row = agg_start_row
    first_data_row = agg_start_row - len(employees)

    # rb. column is at 0, Prezime i ime at 1. So 'Ukupno' should go under the 'Prezime i ime' column
    worksheet.write(total_row, 0, '', formats['header_format'])
    worksheet.write(total_row, 1, 'Ukupno', formats['header_format'])

    total_columns = range(2, col_offset + 2)
    for col_num in total_columns:
        column_letter_value = xlsxwriter.utility.xl_col_to_name(col_num)
        formula = f"=SUM({column_letter_value}{first_data_row + 1}:{column_letter_value}{total_row})"
        worksheet.write_formula(total_row, col_num, formula, formats['total_format'])

def generate_excess_hours(worksheet, employees, current_year, current_month, formats):
    worksheet.set_column(0, 0, 5)  # rb. column
    worksheet.set_column(1, 1, 30)  # Employee name column
    worksheet.set_column(2, 4, 20)

    agg_start_row = 0
    worksheet.merge_range(agg_start_row, 0, agg_start_row, 4, 'EVIDENCIJA VIŠKA-MANJKA SATI', formats['title_format'])
    agg_start_row += 1

    # Headers: rb. first, then Prezime i ime
    sub_headers = [
        'rb.', 'PREZIME I IME',
        f"Fond sati s {monthrange(current_year, current_month - 1 if current_month > 1 else 12)[1]}.{(current_month - 1 if current_month > 1 else 12)}",
        f"Višak fonda ({current_month} {current_year})",
        f"Fond sati s {monthrange(current_year, current_month)[1]}.{current_month}"
    ]

    for i, header in enumerate(sub_headers):
        worksheet.merge_range(agg_start_row, i, agg_start_row + 1, i, header, formats['header_format'])

    agg_start_row += 2

    for idx, employee in enumerate(employees):
        try:
            prev_year = current_year if current_month > 1 else current_year - 1
            prev_month = current_month - 1 if current_month > 1 else 12
            previous_excess_record = ExcessHours.objects.get(employee=employee, year=prev_year, month=prev_month)
            previous_excess = previous_excess_record.excess_hours
        except ExcessHours.DoesNotExist:
            previous_excess = 0

        current_excess = employee.calculate_visak_sati(current_month, current_year)
        cumulative_excess_current_month = previous_excess + current_excess

        row_format = formats['gray_format'] if idx % 2 == 0 else formats['white_format']

        worksheet.write(agg_start_row, 0, idx + 1, row_format)  # rb.
        worksheet.write(agg_start_row, 1, f"{employee.surname} {employee.name}", row_format) # Prezime i ime
        worksheet.write(agg_start_row, 2, previous_excess, formats['hours_format'])
        worksheet.write(agg_start_row, 3, current_excess, formats['hours_format'])
        worksheet.write(agg_start_row, 4, cumulative_excess_current_month, formats['hours_format'])

        agg_start_row += 1

    total_row = agg_start_row
    first_data_row = agg_start_row - len(employees)

    worksheet.write(total_row, 1, 'Ukupno', formats['header_format'])
    worksheet.write(total_row, 0, '', formats['header_format'])  # Empty cell under 'Prezime i ime'

    total_columns = range(2, 5)  # Columns for numerical data
    for col_num in total_columns:
        column_letter_value = xlsxwriter.utility.xl_col_to_name(col_num)
        formula = f"=SUM({column_letter_value}{first_data_row + 1}:{column_letter_value}{total_row})"
        worksheet.write_formula(total_row, col_num, formula, formats['total_format'])


def generate_vacation_hours(worksheet, employees, current_year, current_month, formats):
    worksheet.set_column(0, 0, 5)  # rb. column
    worksheet.set_column(1, 1, 30)  # Employee name column
    worksheet.set_column(2, 5, 15)

    agg_vacation_start_row = 0
    worksheet.merge_range(agg_vacation_start_row, 0, agg_vacation_start_row, 5, 'EVIDENCIJA GODIŠNJIH ODMORA', formats['title_format'])
    agg_vacation_start_row += 1

    vacation_headers = [
        'rb.', 'Prezime i ime',
        f"GO s {monthrange(current_year, current_month - 1 if current_month > 1 else 12)[1]}.{(current_month - 1 if current_month > 1 else 12)}.{current_year if current_month > 1 else current_year - 1}",
        f"GO sati ({current_month}.{current_year})",
        f"GO dani ({current_month}.{current_year})",
        f"GO s {monthrange(current_year, current_month)[1]}.{current_month}"
    ]

    for i, header in enumerate(vacation_headers):
        worksheet.merge_range(agg_vacation_start_row, i, agg_vacation_start_row + 1, i, header, formats['header_format'])

    agg_vacation_start_row += 2

    for idx, employee in enumerate(employees):
        previous_vacation_record = ExcessHours.objects.filter(
            employee=employee,
            year=current_year if current_month > 1 else current_year - 1,
            month=current_month - 1 if current_month > 1 else 12
        ).first()
        previous_vacation_hours = previous_vacation_record.vacation_hours_used if previous_vacation_record else 0

        current_vacation_hours = WorkDay.objects.filter(
            employee=employee, date__year=current_year, date__month=current_month
        ).aggregate(Sum('vacation_hours'))['vacation_hours__sum'] or 0
        current_vacation_days = current_vacation_hours / 8

        new_cumulative_vacation_hours = previous_vacation_hours + current_vacation_hours
        new_cumulative_vacation_days = new_cumulative_vacation_hours / 8

        row_format = formats['gray_format'] if idx % 2 == 0 else formats['white_format']

        worksheet.write(agg_vacation_start_row, 0, idx + 1, row_format)  # rb.
        worksheet.write(agg_vacation_start_row, 1, f"{employee.surname} {employee.name}", row_format)
        worksheet.write(agg_vacation_start_row, 2, previous_vacation_hours / 8, formats['hours_format'])
        worksheet.write(agg_vacation_start_row, 3, current_vacation_hours, formats['hours_format'])
        worksheet.write(agg_vacation_start_row, 4, current_vacation_days, formats['hours_format'])
        worksheet.write(agg_vacation_start_row, 5, new_cumulative_vacation_days, formats['hours_format'])

        agg_vacation_start_row += 1

    total_row = agg_vacation_start_row
    first_data_row = agg_vacation_start_row - len(employees)

    worksheet.write(total_row, 1, 'Ukupno', formats['header_format'])
    worksheet.write(total_row, 0, '', formats['header_format'])

    # Summations for columns 2 to 5
    for col_num in range(2, 6):
        column_letter_value = xlsxwriter.utility.xl_col_to_name(col_num)
        formula = f"=SUM({column_letter_value}{first_data_row + 1}:{column_letter_value}{total_row})"
        worksheet.write_formula(total_row, col_num, formula, formats['total_format'])




def generate_overtime_hours(worksheet, employees, current_year, current_month, formats):
    worksheet.set_column(0, 0, 5)  # rb. column
    worksheet.set_column(1, 1, 30)  # Employee name column
    worksheet.set_column(2, 8, 20)
    agg_overtime_start_row = 0

    overtime_headers = [
        'rb.', 'Prezime i ime',
        'Prek. pripreme', 'Prek. USLUGA priprema', 'Prek. višak fonda',
        'Prekovremeno slobodan dan', 'Prekovremeno slobodan dan usluga',
        'Ukupno prek.', 'Ukupno prek. USLUGA'
    ]

    worksheet.merge_range(
        agg_overtime_start_row, 0, agg_overtime_start_row, 
        9, 'EVIDENCIJA PREKOVREMENIH SATI', formats['title_format']
    )
    for i, header in enumerate(overtime_headers):
        worksheet.merge_range(agg_overtime_start_row + 1, i, agg_overtime_start_row + 2, i, header, formats['header_format'])
    agg_overtime_start_row += 3

    for idx, employee in enumerate(employees):
        work_days = WorkDay.objects.filter(employee=employee, date__year=current_year, date__month=current_month)
        
        total_overtime_preparation = sum_decimal_hours([day.overtime_hours for day in work_days])
        total_overtime_service = sum_decimal_hours([day.overtime_service for day in work_days])
        total_overtime_excess_fond = sum_decimal_hours([day.overtime_excess_fond for day in work_days])
        total_overtime_free_day = sum_decimal_hours([day.overtime_free_day for day in work_days])
        total_overtime_free_day_service = sum_decimal_hours([day.overtime_free_day_service for day in work_days])

        total_overtime = total_overtime_preparation + total_overtime_service + total_overtime_excess_fond
        total_overtime_service_total = total_overtime_free_day + total_overtime_free_day_service

        row_format = formats['gray_format'] if idx % 2 == 0 else formats['white_format']

        worksheet.write(agg_overtime_start_row, 0, idx + 1, row_format)  # rb.
        worksheet.write(agg_overtime_start_row, 1, f"{employee.surname} {employee.name}", row_format) # Prezime i ime
        worksheet.write(agg_overtime_start_row, 2, total_overtime_preparation, formats['hours_format'])
        worksheet.write(agg_overtime_start_row, 3, total_overtime_service, formats['hours_format'])
        worksheet.write(agg_overtime_start_row, 4, total_overtime_excess_fond, formats['hours_format'])
        worksheet.write(agg_overtime_start_row, 5, total_overtime_free_day, formats['hours_format'])
        worksheet.write(agg_overtime_start_row, 6, total_overtime_free_day_service, formats['hours_format'])
        worksheet.write(agg_overtime_start_row, 7, total_overtime, formats['hours_format'])
        worksheet.write(agg_overtime_start_row, 8, total_overtime_service_total, formats['hours_format'])

        agg_overtime_start_row += 1

    total_row = agg_overtime_start_row
    first_data_row = agg_overtime_start_row - len(employees)

    worksheet.write(total_row, 1, 'Ukupno', formats['header_format'])
    worksheet.write(total_row, 0, '', formats['header_format'])

    total_columns = range(2, 9)  # Columns with numerical data (columns 2 to 8)
    for col_num in total_columns:
        column_letter_value = xlsxwriter.utility.xl_col_to_name(col_num)
        formula_range = f"{column_letter_value}{first_data_row + 1}:{column_letter_value}{total_row}"
        worksheet.write_formula(total_row, col_num, f"=SUM({formula_range})", formats['total_format'])

def sum_decimal_hours(decimal_hours):
    """Sums a list of decimal hours, converting minute values correctly."""
    total_hours = 0
    total_minutes = 0
    for decimal in decimal_hours:
        hours = int(decimal)
        minutes = (decimal - hours) * 100
        total_hours += hours
        total_minutes += minutes

    # Convert total minutes to hours and minutes
    total_hours += total_minutes // 60
    total_minutes = total_minutes % 60

    return total_hours + total_minutes / 100.0

def column_letter(index):
    index -= 1
    result = ''
    while index >= 0:
        result = chr(index % 26 + 65) + result
        index = index // 26 - 1
    return result
