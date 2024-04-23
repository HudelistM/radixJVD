#Django db importovi
from django.db import transaction
from django.db.models import Q
from django.db.models import F
#Import formi i modela
from .forms import UserRegisterForm, EmployeeForm
from .models import ScheduleEntry, ShiftType, Employee
#Django.views importovi
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
#Django http importovi
from django.http import JsonResponse
from django.http import HttpResponse

from django.shortcuts import render, redirect
from django.contrib.auth import login
from datetime import date, timedelta, datetime
from django.contrib import messages
import json
from django.core import serializers
from django.template.loader import render_to_string
from django.shortcuts import get_object_or_404

#Excel importovi
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import CellRichText,TextBlock
from openpyxl.worksheet.page import PageMargins

from django.http import HttpResponse
from io import BytesIO
from openpyxl.styles import Color
from openpyxl.workbook import Workbook
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

import logging

logger = logging.getLogger(__name__)


def is_ajax(request):
    return request.headers.get('x-requested-with') == 'XMLHttpRequest'

def landingPage(request):
    return render(request, "scheduler/landing_page.html")
  
def register(request):
    if request.method == 'POST':
        form = UserRegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)  # Log the user in
            return redirect('landingPage')
    else:
        form = UserRegisterForm()
    return render(request, 'scheduler/register.html', {'form': form})

def get_week_dates(start_date):
    # start_date is assumed to be a Monday
    return [start_date + timedelta(days=i) for i in range(7)]

def radnici(request):
    
    employees = Employee.objects.all()

    if request.method == 'POST':
        form = UserRegisterForm(request.POST)
        if form.is_valid():  
            user = form.save()
            login(request, user) 
            return redirect('landingPage')  
    else:
        form = UserRegisterForm()
    context = {
        'form': form,
        'employees': employees,
    }
    return render(request, 'scheduler/radnici.html', context)


def add_employee(request):
    if request.method == 'POST':
        form = EmployeeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Radnik je uspješno dodan!')
            return redirect('radnici') 
    else:
        form = EmployeeForm()
    
    return render(request, 'scheduler/radnici.html', {'form': form})


def schedule_view(request):
    # Default start_date to the current week's Monday if no parameter is passed
    start_date_str = request.GET.get('week_start')
    if start_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    else:
        today = date.today()
        start_date = today - timedelta(days=today.weekday())  # Ensure we start on Monday
    
    week_dates = get_week_dates(start_date)

    shift_types = ShiftType.objects.all()

    # Construct the schedule data structure
    # We'll use the date as a string and the shift_type ID to identify each cell
    schedule_data = {}
    for day in week_dates:
        day_key = day.strftime('%Y-%m-%d')
        schedule_data[day_key] = {}
        for shift_type in shift_types:
            # Get the first schedule entry for this day and shift_type or None if it doesn't exist
            entry = ScheduleEntry.objects.filter(date=day, shift_type=shift_type).first()
            schedule_data[day_key][shift_type.id] = entry
    #print(schedule_data)

    context = {
        'week_dates': week_dates,
        'shift_types': shift_types,
        'schedule_data': schedule_data,
        'employees': Employee.objects.all(),
    }
    if is_ajax(request):
        # Render only the schedule part of the page for AJAX requests
        schedule_grid_html = render_to_string('scheduler/scheduler_grid_inner.html', context, request)
        return HttpResponse(schedule_grid_html)

    return render(request, 'scheduler/schedule_grid.html', context)

# New view function to serve schedule data as JSON
@require_http_methods(["GET"])
def api_schedule_data(request):
    # Extract query parameters for week start and end dates
    week_start = request.GET.get('week_start')
    week_end = request.GET.get('week_end')
    
    # Parse the dates from strings to datetime objects
    week_start_date = datetime.strptime(week_start, '%Y-%m-%d').date() if week_start else None
    week_end_date = datetime.strptime(week_end, '%Y-%m-%d').date() if week_end else None
    
    if not week_start_date or not week_end_date:
        return JsonResponse({'error': 'Invalid or missing date parameters'}, status=400)
    
    schedule_entries = ScheduleEntry.objects.filter(
        date__range=[week_start_date, week_end_date]
    ).select_related('shift_type').prefetch_related('employees')
    
    schedule_list = []
    for entry in schedule_entries:
        schedule_list.append({
            "date": entry.date.strftime('%Y-%m-%d'),
            "shift_type_id": entry.shift_type.id,
            "employees": list(entry.employees.values('id', 'name', 'surname', 'group'))
        })
    
    return JsonResponse(schedule_list, safe=False)



@csrf_exempt
@require_http_methods(["POST"])
def update_schedule(request):
    try:
        data = json.loads(request.body.decode('utf-8'))
        logger.debug(f"Received data for update: {data}")

        if any(key not in data for key in ['action', 'employeeId', 'shiftTypeId', 'date']):
            return JsonResponse({'error': 'Missing required parameters'}, status=400)

        date = datetime.strptime(data['date'], '%Y-%m-%d').date()
        employee = get_object_or_404(Employee, pk=data['employeeId'])
        shift_type = get_object_or_404(ShiftType, pk=data['shiftTypeId'])

        if data['action'] == 'move':
            original_date = datetime.strptime(data['originalDate'], '%Y-%m-%d').date()
            original_shift_type = get_object_or_404(ShiftType, pk=data['originalShiftTypeId'])

            with transaction.atomic():
                # First, remove the employee from the original entry
                original_entry = ScheduleEntry.objects.filter(date=original_date, shift_type=original_shift_type).first()
                if original_entry:
                    original_entry.employees.remove(employee)
                    # If no employees are left in the schedule entry, delete it
                    if not original_entry.employees.exists():
                        original_entry.delete()

                # Next, add the employee to the new entry, creating if necessary
                schedule_entry, created = ScheduleEntry.objects.get_or_create(
                    date=date, shift_type=shift_type
                )
                schedule_entry.employees.add(employee)

        elif data['action'] == 'add':
            with transaction.atomic():
                schedule_entry, created = ScheduleEntry.objects.get_or_create(
                    date=date, shift_type=shift_type
                )
                schedule_entry.employees.add(employee)

        return JsonResponse({'status': 'success', 'message': 'Schedule updated', 'created': created}, status=200)

    except Exception as e:
        logger.error(f"Exception in update_schedule: {str(e)}")
        return JsonResponse({'error': 'Server error', 'details': str(e)}, status=500)

GROUP_COLOR_MAPPING = {
    1: 'FFFF0000',  # Red
    2: 'FF00FF00',  # Green
    3: 'FF0000FF',  # Blue
    4: 'FFFFFF00',  # Yellow
    # Add more mappings as needed
}

def create_schedule_excel(week_dates, schedule_data, user_name,group_color_mapping):
    # Initialize workbook and sheet
    wb = Workbook()
    ws = wb.active

    # Set page setup for A4 size and orientation
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_margins = PageMargins(left=0.7, right=0.7, top=0.75, bottom=0.75, header=0.3, footer=0.3)


    # Set titles and headers
    title = f"Weekly Schedule Prepared by {user_name}"
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws.append(["Date"] + [d.strftime("%A %d/%m/%Y") for d in week_dates])  # Headers for dates

    # Define some styles for formatting
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="00FFFF00")
    for cell in ws[2]:
        cell.font = header_font
        cell.fill = header_fill

    def create_rich_text_cell(entries, group_color_mapping):
        rich_text = CellRichText()
        for index, entry in enumerate(entries):
            # Get employee name and group color
            name = f"{entry.name} {entry.surname}"
            group = int(entry.group)
            group_color = group_color_mapping.get(group, '00000000')  # Default to black if group not in mapping

            # If the group has a specific color, create TextBlock with InlineFont
            if group_color != '00000000':  
                font = Font(color=group_color, bold=True, size=12)
                inline_font = InlineFont(font)
                text_block = TextBlock(inline_font, name)
                rich_text.append(text_block)
            else:  
                # If no specific color, append the name as plain text
                rich_text.append(name)

            # Add a newline if there are more entries to add
            if index < len(entries) - 1:
                rich_text.append("\n")

        return rich_text

    # Populate the schedule
    for shift_type in ShiftType.objects.all():
        row = [shift_type.name]
        for date in week_dates:
            day_key = date.strftime('%Y-%m-%d')
            entry = schedule_data[day_key].get(shift_type.id)
            if entry:
                # Get employees for the schedule entry
                employees = entry.employees.all() if hasattr(entry, 'employees') else []
                cell_value = create_rich_text_cell(employees,group_color_mapping)
                row.append(cell_value)
            else:
                row.append(None)
        ws.append(row)

    # Adjust column widths
    for column_cells in ws.columns:
        length = max(len(as_text(cell.value)) for cell in column_cells if cell.value)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length

    return wb

def as_text(value):
    if value is None:
        return ""
    if isinstance(value, CellRichText):
        return str(value)
    return str(value)


def download_schedule(request):
    # Generate dates for the desired schedule week
    week_dates = get_week_dates(date.today())  # Replace with the dates you want to include in your Excel

    # Gather your schedule data here
        # Default start_date to the current week's Monday if no parameter is passed
    start_date_str = request.GET.get('week_start')
    if start_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    else:
        today = date.today()
        start_date = today - timedelta(days=today.weekday())  # Ensure we start on Monday
    
    week_dates = get_week_dates(start_date)

    shift_types = ShiftType.objects.all()

    # Construct the schedule data structure
    # We'll use the date as a string and the shift_type ID to identify each cell
    schedule_data = {}
    for day in week_dates:
        day_key = day.strftime('%Y-%m-%d')
        schedule_data[day_key] = {}
        for shift_type in shift_types:
            # Get the first schedule entry for this day and shift_type or None if it doesn't exist
            entry = ScheduleEntry.objects.filter(date=day, shift_type=shift_type).first()
            schedule_data[day_key][shift_type.id] = entry

    # Create an Excel workbook with the schedule data
    current_user_name = request.user.get_full_name() if request.user.is_authenticated else "Unknown User"
    wb = create_schedule_excel(week_dates, schedule_data, current_user_name,GROUP_COLOR_MAPPING)

    # Save the workbook to a BytesIO object
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Construct response
    response = HttpResponse(content=excel_file.read())
    response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response['Content-Disposition'] = 'attachment; filename=Raspored.xlsx'
    return response